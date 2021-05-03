from openpyxl import load_workbook
import re
import yaml
from unidecode import unidecode
import random
import csv
import chardet


def get_file_encoding(src_file_path):
    """
    Get the encoding type of a file
    :param src_file_path: file path
    :return: str - file encoding type
    """

    with open(src_file_path) as src_file:
        return src_file.encoding


def get_file_encoding_chardet(file_path):
    """
    Get the encoding of a file using chardet package
    :param file_path:
    :return:
    """
    with open(file_path, 'rb') as f:
        result = chardet.detect(f.read())
        return result['encoding']


with open("./config.yaml") as f:
    config = yaml.load(f, Loader=yaml.FullLoader)


def compose_username(f_name, l_name, etablissment, u_names):
    f_name = re.sub("[',\"]", '', f_name.lower())
    l_name = re.sub("[',\"]", '', l_name.lower())
    etablissment = re.sub("[',\"]", '', etablissment.lower())
    gen = unidecode("{}_{}".format(
        f_name.split()[0].replace('-', '_'),
        "_".join(etablissment.split()[0:2]).replace('-', '_').replace('.', '_')
    ))
    copy = gen
    counter = 1
    # eliminate duplicate username(s)
    while gen in u_names:
        gen = "{}_{}".format(copy, counter)
        counter += 1

    return gen


def compose_last_name(f_name):
    split = [name.capitalize() for name in f_name.split()]
    if len(split) > 1:
        return " ".join(split[0:1]), " ".join(split[1:])
    else:
        return split[0], split[0]


def main():
    wb = load_workbook(config['FILENAME'])
    usernames = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        labels = [[field.value.lower() for field in row if field.value] for row in ws.iter_rows(max_row=1)][0]

        USERNAME_COL = labels.index('username') + 1
        F_NAME_COL = labels.index('firstname') + 1
        L_NAME_COL = labels.index('lastname') + 1
        PASS_COL = labels.index('password') + 1
        ETABLISSMENT_COL = labels.index('profile_field_etablissement') + 1

        current_row = 1
        for row in ws.iter_rows():
            col = 1
            for cell in row:
                if cell.value in labels:
                    break
                if col == USERNAME_COL:  # Stop at Username Column
                    first_name = ws.cell(row=current_row, column=F_NAME_COL).value
                    last_name = ws.cell(row=current_row, column=L_NAME_COL).value
                    etabliss = ws.cell(row=current_row, column=ETABLISSMENT_COL).value
                    password = ws.cell(row=current_row, column=PASS_COL).value
                    if etabliss is None:
                        continue
                    else:
                        etabliss = re.sub(' +', ' ', etabliss)

                    if first_name is None and last_name is None:
                        continue
                    else:
                        if first_name is not None:
                            first_name = re.sub(' +', ' ', first_name)
                        if last_name is not None:
                            last_name = re.sub(' +', ' ', last_name)

                    if config['UPDATE_NAMES']:
                        if last_name is None:
                            first_name, last_name = compose_last_name(first_name)
                            ws.cell(row=current_row, column=F_NAME_COL, value=first_name)
                            ws.cell(row=current_row, column=L_NAME_COL, value=last_name)
                        if first_name is None:
                            first_name, last_name = compose_last_name(last_name)
                            ws.cell(row=current_row, column=F_NAME_COL, value=first_name)
                            ws.cell(row=current_row, column=L_NAME_COL, value=last_name)
                    if config['UPDATE_PASSWORD'] or type(password) == int:
                        # if password is None or type(password) == int:
                        pass_gen = chr(random.randrange(ord('a'), ord('z')))
                        ws.cell(row=current_row, column=PASS_COL, value=pass_gen)

                    if config['UPDATE_USERNAME']:
                        username = compose_username(first_name, last_name, etabliss, usernames)

                        usernames.append(username)
                        ws.cell(row=cell.row, column=cell.column, value=username)
                col += 1
            current_row += 1
        if config['EXPORT_CSV']:
            CSV = "{}/{}. {}.csv".format(config['SAVE_PATH'], wb.sheetnames.index(sheet_name) + 1, ws.title)
            with open(CSV, 'w', newline="", encoding='utf-8') as fh:
                c = csv.writer(fh, delimiter=config['CSV_DELIMITER'])
                for r in ws.rows:
                    if r[0].value is not None:  # ignore empty fields.
                        c.writerow([cell.value for cell in r])
            print('Endcoding: ' + str(get_file_encoding(CSV)))
            print('Endcoding with chardet: ' + str(get_file_encoding_chardet(CSV)))
    wb.save("{}/{}.xlsx".format(config['SAVE_PATH'], 'final'))  # , ntpath.basename(config['FILENAME'])

    print("ALL DONE!!!")


if __name__ == '__main__':
    main()
